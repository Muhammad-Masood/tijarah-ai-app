from openpyxl import load_workbook
ws = load_workbook(r"c:\Users\dossani\Downloads\Acceleration Cohort 4- Application Deliverables - Tijarah AI_filled.xlsx")["Product listed"]
print("merged:", [str(r) for r in ws.merged_cells.ranges])
for r in range(1, 21):
    row = [(c, ws.cell(r,c).value) for c in range(1, 6) if ws.cell(r,c).value is not None]
    if row:
        print(r, row)
