from openpyxl import load_workbook
PATH = r"c:\Users\dossani\Downloads\Acceleration Cohort 4- Application Deliverables - Tijarah AI_filled.xlsx"
ws = load_workbook(PATH)["Information Memorandum"]
for r in range(22, 35):
    row = {c: ws.cell(r,c).value for c in range(2, 13) if ws.cell(r,c).value}
    if row:
        print(r, row)
print("--- impact ---")
for r in range(54, 62):
    row = {c: ws.cell(r,c).value for c in range(13, 17) if ws.cell(r,c).value}
    if row:
        print(r, row)
