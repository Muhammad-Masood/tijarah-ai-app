"""Audit NIC filled Excel workbook."""
import sys
from openpyxl import load_workbook

PATH = r"c:\Users\dossani\Downloads\Acceleration Cohort 4- Application Deliverables - Tijarah AI_filled.xlsx"
OUT = r"C:\Users\dossani\tijarah-ai-app\scripts\audit_excel_output.txt"

wb = load_workbook(PATH, data_only=True)
lines = []

for name in wb.sheetnames:
    ws = wb[name]
    lines.append(f"\n{'='*80}\nSHEET: {name}\n{'='*80}")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        cells = []
        for c in row:
            v = c.value
            if v is None:
                continue
            s = str(v).strip()
            if not s or s in ("Response:", "Response:\n", "Response: \n", "NaN"):
                continue
            if len(s) > 120 and ("eg." in s.lower() or "example" in s.lower()):
                continue
            cells.append(f"{c.coordinate}={s[:300]}")
        if cells:
            lines.append(" | ".join(cells))

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))

print(f"Wrote {len(lines)} lines to {OUT}")
