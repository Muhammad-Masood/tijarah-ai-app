"""Find top-left cells for writing."""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PATH = r"c:\Users\dossani\Downloads\Acceleration Cohort 4- Application Deliverables - Tijarah AI_filled.xlsx"
wb = load_workbook(PATH)
ws = wb["Information Memorandum"]

def is_merged(coord):
    for r in ws.merged_cells.ranges:
        if coord in r:
            return str(r)
    return None

for addr in ["A5", "B5", "B55", "B56", "B57", "B58", "B59", "G24", "G34", "F57", "N39", "B2", "B18"]:
    c = ws[addr]
    print(addr, "merged:", is_merged(addr), "value:", c.value)
