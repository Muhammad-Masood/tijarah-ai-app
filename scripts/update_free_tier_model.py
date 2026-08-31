"""Recalculate and apply free-tier-no-AI financial model."""
from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

PATHS = [
    r"c:\Users\dossani\Downloads\Acceleration Cohort 4- Application Deliverables - Tijarah AI_filled.xlsx",
    r"C:\Users\dossani\tijarah-ai-app\docs\NIC-ACCELERATION-COHORT-4-Application-Deliverables-Tijarah-AI.xlsx",
]
FX = 280
AI_USD_PER_PAID = 4.90
HOSTING_USD = 70
PLATFORM_FEE_USD_M1 = 143

Y1_PAID = [5, 10, 17, 25, 34, 45, 58, 74, 92, 113, 138, 168]
Y1_MRR_PKR = [31800, 69324, 113602, 165851, 227504, 300255, 386100, 487398, 606930, 747978, 914414, 1110808]
NEW_SIGNUPS_Y1 = [80, 94, 111, 131, 155, 183, 216, 255, 301, 355, 419, 494]

Y1_REV_USD = 18436
Y2_REV_USD = 46089
Y3_REV_USD = 82960
Y1_REV_PKR_MN = 5.16408
Y2_REV_PKR_MN = 12.9102
Y3_REV_PKR_MN = 23.2288


def cogs_usd_paid_only(paid: int, month_index: int) -> float:
    cost = paid * AI_USD_PER_PAID + HOSTING_USD
    if month_index == 0:
        cost += PLATFORM_FEE_USD_M1
    return cost


def y2_paid_series() -> list[int]:
    merchants = []
    c = 2794
    for _ in range(12):
        c = int(c * 1.12)
        merchants.append(c)
    return [int(c * 0.06) for c in merchants]


def y2_mrr_series() -> list[int]:
    m = 1110808
    vals = []
    for _ in range(12):
        m = int(m * 1.095)
        vals.append(m)
    scale = (Y2_REV_PKR_MN * 1_000_000) / sum(vals)
    return [int(v * scale) for v in vals]


def annual_cogs_usd(paid_monthly: list[int]) -> float:
    return sum(cogs_usd_paid_only(p, i) for i, p in enumerate(paid_monthly))


Y1_COGS_USD = annual_cogs_usd(Y1_PAID)
Y2_COGS_USD = annual_cogs_usd(y2_paid_series())
Y3_PAID_AVG = int((y2_paid_series()[-1] + y2_paid_series()[-1] * 1.5) / 2)  # ~978 avg
Y3_COGS_USD = Y3_PAID_AVG * AI_USD_PER_PAID * 12 + HOSTING_USD * 12

Y1_GP_USD = Y1_REV_USD - Y1_COGS_USD
Y2_GP_USD = Y2_REV_USD - Y2_COGS_USD
Y3_GP_USD = Y3_REV_USD - Y3_COGS_USD

Y1_GP_PKR_MN = round(Y1_GP_USD * FX / 1_000_000, 4)
Y2_GP_PKR_MN = round(Y2_GP_USD * FX / 1_000_000, 4)
Y3_GP_PKR_MN = round(Y3_GP_USD * FX / 1_000_000, 4)

GROSS_MARGIN_Y1 = round(Y1_GP_USD / Y1_REV_USD * 100)
GROSS_MARGIN_Y2 = round(Y2_GP_USD / Y2_REV_USD * 100)
GROSS_MARGIN_Y3 = round(Y3_GP_USD / Y3_REV_USD * 100)


def set_cell(ws, coord: str, value) -> None:
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        for merged in list(ws.merged_cells.ranges):
            if coord in merged:
                ws.unmerge_cells(str(merged))
                break
        cell = ws[coord]
    cell.value = value


def apply(path: str) -> None:
    y2_paid = y2_paid_series()
    y2_mrr = y2_mrr_series()
    wb = load_workbook(path)
    fm = wb["Financial Model"]

    fm["F4"] = "AI inference billed to paid tiers only; free = sync/listing/compute (no LLM)"
    fm["C4"] = "New signups / month (Y1)"
    fm["D4"] = "Paid subscribers (AI access)"
    fm["E4"] = "MRR (PKR)"

    for i in range(12):
        r = 5 + i
        paid = Y1_PAID[i]
        mrr = Y1_MRR_PKR[i]
        cogs_usd = cogs_usd_paid_only(paid, i)
        cogs_pkr = int(cogs_usd * FX)
        gp_pkr = mrr - cogs_pkr

        fm[f"B{r}"] = 0.18 if i > 0 else None
        fm[f"C{r}"] = NEW_SIGNUPS_Y1[i]
        fm[f"D{r}"] = paid
        fm[f"E{r}"] = mrr
        fm[f"I{r}"] = mrr
        fm[f"L{r}"] = 0
        fm[f"O{r}"] = gp_pkr
        fm[f"P{r}"] = gp_pkr

        er = 36 + i
        fm[f"B{er}"] = 2
        fm[f"L{er}"] = int(HOSTING_USD * FX)
        fm[f"H{er}"] = int(paid * AI_USD_PER_PAID * FX)
        fm[f"K{er}"] = int(PLATFORM_FEE_USD_M1 * FX) if i == 0 else 0

    fm["F18"] = "Y2 — AI COGS on paid subscribers only"

    for i in range(12):
        r = 19 + i
        paid = y2_paid[i]
        mrr = y2_mrr[i]
        cogs_pkr = int(cogs_usd_paid_only(paid, i) * FX)
        gp_pkr = mrr - cogs_pkr
        fm[f"B{r}"] = 0.095
        fm[f"C{r}"] = int(2794 * (1.12 ** (i + 1)))
        fm[f"D{r}"] = paid
        fm[f"E{r}"] = mrr
        fm[f"I{r}"] = mrr
        fm[f"O{r}"] = gp_pkr
        fm[f"P{r}"] = gp_pkr

        er = 50 + i
        fm[f"B{er}"] = 3
        fm[f"L{er}"] = int(HOSTING_USD * FX)
        fm[f"H{er}"] = int(paid * AI_USD_PER_PAID * FX)
        fm[f"F{er}"] = 150000 if i >= 3 else 0

    im = wb["Information Memorandum"]
    set_cell(
        im,
        "B32",
        "FX PKR 280/USD; Rs 5K/7.5K/10K pricing; 6% free→paid; 18% MoM signup growth; "
        "AI (~$4.90/paid merchant/mo) on paid tiers ONLY — free tier is non-AI sync, listing, and computed metrics.",
    )
    set_cell(
        im,
        "B24",
        "Monthly subscription: Free (non-AI) / Starter Rs 5,000 / Growth Rs 7,500 / Enterprise Rs 10,000.",
    )
    set_cell(im, "N39", "Free")
    set_cell(im, "O39", 0)
    set_cell(im, "N40", "Starter")
    set_cell(im, "O40", 5000)
    set_cell(im, "N41", "Growth")
    set_cell(im, "O41", 7500)
    # Row 42 enterprise if exists
    set_cell(im, "N42", "Enterprise")
    set_cell(im, "O42", 10000)

    for col, margin in zip(["F", "G", "H"], [GROSS_MARGIN_Y1, GROSS_MARGIN_Y2, GROSS_MARGIN_Y3]):
        set_cell(im, f"{col}57", margin)
        set_cell(im, f"{col}58", margin)
        set_cell(im, f"{col}59", margin)

    set_cell(im, "F67", Y1_GP_PKR_MN)
    set_cell(im, "G67", Y2_GP_PKR_MN)
    set_cell(im, "H67", Y3_GP_PKR_MN)
    set_cell(im, "F68", Y1_GP_PKR_MN)
    set_cell(im, "G68", Y2_GP_PKR_MN)
    set_cell(im, "H68", Y3_GP_PKR_MN)

    set_cell(
        im,
        "K27",
        "Free tier has no AI cost; paid upgrade unlocks AI agents — clear value gap and pricing validation.",
    )

    bp = wb["Business Plan"]
    set_cell(
        bp,
        "D8",
        f"Modelled Y1 revenue ${Y1_REV_USD:,} with ~{GROSS_MARGIN_Y1}% gross margin (AI costs apply to paid tiers only; "
        f"free tier = store sync, cross-platform listing, and computed metrics without LLM). Y2 ${Y2_REV_USD:,}, Y3 ${Y3_REV_USD:,}.",
    )

    lbc = wb["Lean Business Canvas"]
    set_cell(
        lbc,
        "K11",
        "Freemium: Free (non-AI sync/listing/compute) / Starter Rs 5,000 / Growth Rs 7,500 / Enterprise Rs 10,000 per month.",
    )
    set_cell(
        lbc,
        "A11",
        "AI inference on paid tiers only (~$4.90/paid merchant/mo); hosting ~$70/mo; app-store fees; future team/marketing.",
    )

    wb.save(path)
    print(f"Updated {path}")


if __name__ == "__main__":
    print("Y1 COGS USD:", round(Y1_COGS_USD, 2), "GP USD:", round(Y1_GP_USD, 2), "Margin:", GROSS_MARGIN_Y1)
    print("Y2 COGS USD:", round(Y2_COGS_USD, 2), "GP USD:", round(Y2_GP_USD, 2), "Margin:", GROSS_MARGIN_Y2)
    print("Y3 COGS USD:", round(Y3_COGS_USD, 2), "GP USD:", round(Y3_GP_USD, 2), "Margin:", GROSS_MARGIN_Y3)
    print("GP PKR Mn:", Y1_GP_PKR_MN, Y2_GP_PKR_MN, Y3_GP_PKR_MN)
    for path in PATHS:
        apply(path)
