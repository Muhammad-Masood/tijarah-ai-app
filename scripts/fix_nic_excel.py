"""Fix NIC Tijarah AI filled Excel for submission."""
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

PATH = r"c:\Users\dossani\Downloads\Acceleration Cohort 4- Application Deliverables - Tijarah AI_filled.xlsx"
FX = 280


def set_cell(ws, coord: str, value) -> None:
    """Write to coord, unmerging if the cell is part of a merged range."""
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        for merged in list(ws.merged_cells.ranges):
            if coord in merged:
                ws.unmerge_cells(str(merged))
                break
        cell = ws[coord]
    cell.value = value

Y1 = {
    "cumulative_merchants": [80, 174, 286, 417, 572, 755, 971, 1226, 1527, 1882, 2300, 2794],
    "paid_subscribers": [5, 10, 17, 25, 34, 45, 58, 74, 92, 113, 138, 168],
    "mrr_pkr": [31800, 69324, 113602, 165851, 227504, 300255, 386100, 487398, 606930, 747978, 914414, 1110808],
    "cogs_usd": [605, 925, 1470, 2114, 2874, 3771, 4829, 6078, 7552, 9290, 11342, 13763],
}

GROSS_MARGIN_PCT = -251
new_signups_y1 = [80, 94, 111, 131, 155, 183, 216, 255, 301, 355, 419, 494]

# Year 2 monthly projections
m12 = Y1["mrr_pkr"][-1]
y2_mrr = []
m = m12
for _ in range(12):
    m = int(m * 1.095)
    y2_mrr.append(m)
scale = (12.9102 * 1_000_000) / sum(y2_mrr)
y2_mrr = [int(v * scale) for v in y2_mrr]
y2_cogs_usd_monthly = 161278 / 12

y2_merchants = []
c = Y1["cumulative_merchants"][-1]
for _ in range(12):
    c = int(c * 1.12)
    y2_merchants.append(c)
y2_paid = [int(c * 0.06) for c in y2_merchants]

wb = load_workbook(PATH)

# --- Financial Model ---
fm = wb["Financial Model"]
fm["C4"] = "New signups / month (Y1)"
fm["D4"] = "Paid subscribers"
fm["E4"] = "MRR (PKR)"
fm["F4"] = "6% paid conversion; 18% MoM signup growth; ARPU Rs 6,625"
fm["I4"] = "Monthly revenue (PKR)"
fm["L4"] = "CAC (PKR) — not modelled"
fm["O4"] = "Gross profit (PKR)"
fm["P4"] = "EBIT (PKR) — pre-salary"

for i in range(12):
    r = 5 + i
    mrr = Y1["mrr_pkr"][i]
    cogs_pkr = int(Y1["cogs_usd"][i] * FX)
    gp_pkr = mrr - cogs_pkr

    fm[f"B{r}"] = 0.18 if i > 0 else None
    fm[f"C{r}"] = new_signups_y1[i]
    fm[f"D{r}"] = Y1["paid_subscribers"][i]
    fm[f"E{r}"] = mrr
    fm[f"I{r}"] = mrr
    fm[f"L{r}"] = 0
    fm[f"O{r}"] = gp_pkr
    fm[f"P{r}"] = gp_pkr

    er = 36 + i
    fm[f"B{er}"] = 2
    fm[f"L{er}"] = 19600
    fm[f"H{er}"] = max(cogs_pkr - 19600, 0)
    fm[f"E{er}"] = 0
    fm[f"F{er}"] = 0
    fm[f"J{er}"] = 0
    fm[f"K{er}"] = 40000 if i == 0 else 0

fm["C18"] = "Y2 cumulative merchants"
fm["D18"] = "Y2 paid subscribers"
fm["E18"] = "Y2 MRR (PKR)"
fm["F18"] = "~9.5% MoM MRR growth (modelled)"

for i in range(12):
    r = 19 + i
    mrr = y2_mrr[i]
    cogs_pkr = int(y2_cogs_usd_monthly * FX)
    gp_pkr = mrr - cogs_pkr
    fm[f"B{r}"] = 0.095
    fm[f"C{r}"] = y2_merchants[i]
    fm[f"D{r}"] = y2_paid[i]
    fm[f"E{r}"] = mrr
    fm[f"I{r}"] = mrr
    fm[f"L{r}"] = 0
    fm[f"O{r}"] = gp_pkr
    fm[f"P{r}"] = gp_pkr

    er = 50 + i
    fm[f"B{er}"] = 3
    fm[f"L{er}"] = 19600
    fm[f"H{er}"] = max(cogs_pkr - 19600, 0)
    fm[f"F{er}"] = 150000 if i >= 3 else 0

fm["L35"] = 19600
fm["B35"] = 2

# --- Cap table ---
wb["Pre-Investment CapTable"]["A4"] = "Muhammad (Founder & CEO)"

# --- Information Memorandum ---
im = wb["Information Memorandum"]
set_cell(im, "B3", "https://tijarah-ai-web.vercel.app/")
set_cell(im, "D5", "https://tijarah-ai-web.vercel.app/ (product); NIC pack in project docs/")

# Restore Key Numbers labels + values
key_rows = [
    (55, "Existing Investment (Equity) (PKR)", "Bootstrapped / founder sweat equity"),
    (56, "Investment Sought (PKR)", "Primarily NIC mentorship; open to aligned grant/equity up to PKR 9M"),
    (57, "Owner's contribution (PKR)", "Bootstrapped founder + co-founder time to date"),
    (58, "Business's Own Cashflows (PKR)", "Pre-revenue at scale; no operating cashflows yet"),
    (59, "Total Investment Required (PKR)", 9000000),
    (60, "Financing Type, Collateral", "Grant / equity / accelerator support (if available)"),
    (61, "Interest Rate (%) & Grace period", "N/A — not debt-financed at this stage"),
    (62, "Repayment Duration & Repayment", "N/A"),
    (63, "Mode", "Mentorship + optional aligned capital"),
    (64, "Expected Financing Date", "Aug 2026 (NIC Cohort 4)"),
]
for row, label, value in key_rows:
    set_cell(im, f"B{row}", label)
    set_cell(im, f"C{row}", value)

# Investment utilization (PKR)
set_cell(im, "G24", 2500000)
set_cell(im, "G27", 800000)
set_cell(im, "G28", 500000)
set_cell(im, "G30", 4000000)
set_cell(im, "G33", 1200000)
set_cell(im, "G34", "=SUM(G24,G27,G28,G30,G33)")

# Pricing
set_cell(im, "N39", "Starter")
set_cell(im, "O39", 5000)
set_cell(im, "N40", "Growth")
set_cell(im, "O40", 7500)
set_cell(im, "N41", "Enterprise")
set_cell(im, "O41", 10000)

# Margins & EBITDA (pre-salary; equals gross profit in model)
for col in ["F", "G", "H"]:
    set_cell(im, f"{col}57", GROSS_MARGIN_PCT)
    set_cell(im, f"{col}58", GROSS_MARGIN_PCT)
    set_cell(im, f"{col}59", GROSS_MARGIN_PCT)
set_cell(im, "F68", -12.9302)
set_cell(im, "G68", -32.2529)
set_cell(im, "H68", -58.0334)

# Impact summary
set_cell(im, "N57", 2)
set_cell(im, "O57", 4)
set_cell(im, "P57", 100)
set_cell(im, "N58", 1)
set_cell(im, "O58", 3)
set_cell(im, "P58", 200)
set_cell(im, "M60", "Cloud/AI providers; Daraz; Shopify")
set_cell(im, "M61", "Pakistani e-commerce SMB sellers")

# Product listed
pl = wb["Product listed"]
set_cell(pl, "C2", "Not yet submitted (Q4 2026 target)")
set_cell(pl, "C3", "Not yet submitted (Q4 2026 target)")
set_cell(pl, "C18", "Available on request — private beta / Expo dev build")

cp = wb["Customer Persona Canvas"]
set_cell(cp, "A3", 'Persona: "Ahmed" — online seller, age 25–45, based in Karachi/Lahore/Islamabad')

wb.save(PATH)
print(f"Saved: {PATH}")
print("Y1 M1 gross profit PKR:", fm["O5"].value)
print("Investment total PKR:", im["G34"].value)
