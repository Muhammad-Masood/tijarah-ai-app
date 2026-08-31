"""Inspect full Financial Model sheet."""
from openpyxl import load_workbook

PATH = r"c:\Users\dossani\Downloads\Acceleration Cohort 4- Application Deliverables - Tijarah AI_filled.xlsx"
wb = load_workbook(PATH, data_only=False)
ws = wb["Financial Model"]

lines = []
for row in range(1, ws.max_row + 1):
    parts = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        v = cell.value
        if v is not None and str(v).strip():
            parts.append(f"{cell.column_letter}{row}={v}")
    if parts:
        lines.append(f"R{row}: " + " | ".join(parts))

with open(r"C:\Users\dossani\tijarah-ai-app\scripts\financial_model_full.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print(f"rows with data: {len(lines)}")
