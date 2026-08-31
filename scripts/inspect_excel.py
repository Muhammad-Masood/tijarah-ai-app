"""Inspect Financial Model and Info Memo cell layout."""
from openpyxl import load_workbook

PATH = r"c:\Users\dossani\Downloads\Acceleration Cohort 4- Application Deliverables - Tijarah AI_filled.xlsx"
wb = load_workbook(PATH, data_only=False)

for sheet_name in ["Financial Model", "Information Memorandum", "Pre-Investment CapTable"]:
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} max_row={ws.max_row} max_col={ws.max_column} ===")
    for row in range(1, min(ws.max_row + 1, 75)):
        parts = []
        for col in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row, col)
            v = cell.value
            if v is not None and str(v).strip():
                col_letter = cell.column_letter
                parts.append(f"{col_letter}{row}={v!r:.80}" if len(repr(v)) > 80 else f"{col_letter}{row}={v!r}")
        if parts:
            print(f"R{row}: " + " | ".join(parts))
