from openpyxl import load_workbook
PATH = r"c:\Users\dossani\Downloads\Acceleration Cohort 4- Application Deliverables - Tijarah AI_filled.xlsx"
ws = load_workbook(PATH)["Information Memorandum"]
for r in range(1, 6):
    print(r, [ws.cell(r,c).value for c in range(1, 6)])
print("---")
for r in range(53, 72):
    row = [ws.cell(r,c).value for c in range(1, 8)]
    if any(row):
        print(r, row)
